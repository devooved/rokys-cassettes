import json
from pathlib import Path
from openpyxl import load_workbook


# ==============================
# FILES
# ==============================

BASE_DIR = Path(__file__).parent

EXCEL_FILE = BASE_DIR / "rokys-albums.xlsx"
JSON_FILE = BASE_DIR / "albums.json"
IMAGES_DIR = BASE_DIR / "images"


# ==============================
# LOAD EXISTING JSON
# ==============================

existing_albums = {}

if JSON_FILE.exists():

    with open(
        JSON_FILE,
        "r",
        encoding="utf-8"
    ) as file:

        existing_data = json.load(file)

    if not isinstance(existing_data, list):
        raise ValueError(
            "albums.json must contain a JSON list."
        )

    for album in existing_data:

        if not isinstance(album, dict):
            continue

        number = str(
            album.get("number", "")
        ).strip()

        if number:
            existing_albums[number] = album


# ==============================
# READ EXCEL
# ==============================

workbook = load_workbook(
    EXCEL_FILE,
    data_only=True
)

sheet = workbook.active

headers = [
    cell.value
    for cell in sheet[1]
]

required_headers = [
    "number",
    "artist",
    "title",
]

if headers != required_headers:

    raise ValueError(
        f"Excel headers must be exactly: "
        f"{required_headers}"
    )


albums = []


# ==============================
# VALIDATE + CONVERT
# ==============================

for row_number, row in enumerate(
    sheet.iter_rows(
        min_row=2,
        values_only=True
    ),
    start=2
):

    # ----------------------------------------------------------
    # Skip completely empty rows
    # ----------------------------------------------------------

    if all(
        value is None
        for value in row
    ):

        continue

    number, artist, title = row

    number = (
        str(number).strip()
        if number is not None
        else ""
    )

    artist = (
        str(artist).strip()
        if artist is not None
        else ""
    )

    title = (
        str(title).strip()
        if title is not None
        else ""
    )


    # ----------------------------------------------------------
    # Validate required fields
    # ----------------------------------------------------------

    if not number:

        raise ValueError(
            f"Row {row_number}: missing number"
        )

    if not artist:

        raise ValueError(
            f"Row {row_number}: missing artist"
        )

    if not title:

        raise ValueError(
            f"Row {row_number}: missing title"
        )


    # ----------------------------------------------------------
    # Check image
    # ----------------------------------------------------------

    image_file = (
        IMAGES_DIR /
        f"{number}.webp"
    )

    if not image_file.exists():

        raise ValueError(
            f"Row {row_number}: "
            f"missing image {image_file.name}"
        )


    # ----------------------------------------------------------
    # Check duplicate numbers
    # ----------------------------------------------------------

    if any(
        album["number"] == number
        for album in albums
    ):

        raise ValueError(
            f"Row {row_number}: "
            f"duplicate album number {number}"
        )


    # ----------------------------------------------------------
    # Preserve existing links
    # ----------------------------------------------------------

    existing_album = existing_albums.get(
        number
    )

    if existing_album:

        youtube_url = existing_album.get(
            "youtube",
            ""
        )

        spotify_url = existing_album.get(
            "spotify",
            ""
        )

    else:

        youtube_url = ""
        spotify_url = ""


    # ----------------------------------------------------------
    # Build album
    # ----------------------------------------------------------

    albums.append({
        "number": number,
        "artist": artist,
        "title": title,
        "image": f"images/{number}.webp",
        "youtube": youtube_url,
        "spotify": spotify_url
    })


# ==============================
# WRITE JSON
# ==============================

with open(
    JSON_FILE,
    "w",
    encoding="utf-8"
) as file:

    json.dump(
        albums,
        file,
        ensure_ascii=False,
        indent=4
    )


print(
    f"Successfully converted "
    f"{len(albums)} albums."
)

print(
    f"Created: {JSON_FILE.name}"
)

print(
    "Existing Spotify and YouTube links "
    "were preserved."
)